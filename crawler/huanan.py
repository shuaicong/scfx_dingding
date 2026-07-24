"""华南粮网爬虫

采集华南粮网（gdgrain.com）的文章和附件。
"""
import logging
import re
import os
from datetime import datetime, timezone, timedelta
from typing import Any

import requests
import html2text

logger = logging.getLogger(__name__)

# 北京时间
CST = timezone(timedelta(hours=8))

LIST_URL = "https://www.gdgrain.com/sgtcPortals-front/sgtc/portals/SPql001"
DETAIL_URL = "https://gdgrain.com/sgtcPortals-front/sgtc/portals/SPql002"
BASE_URL = "https://www.gdgrain.com"
ATTACH_DIR = "data/attachments"  # 附件本地缓存目录

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
    "Content-Type": "application/json;charset=utf-8",
    "X-Requested-With": "XMLHttpRequest",
}

# session 复用
_SESSION: requests.Session | None = None


def _get_session() -> requests.Session:
    global _SESSION
    if _SESSION is None:
        s = requests.Session()
        s.get("https://www.gdgrain.com/", headers=HEADERS, timeout=15)
        _SESSION = s
    return _SESSION


def _timestamp_to_str(ts_ms: int) -> str:
    """毫秒时间戳 → 北京时间 2026-07-23 09:34"""
    return (datetime.fromtimestamp(ts_ms / 1000, tz=CST)
            .strftime("%Y-%m-%d %H:%M:%S"))


# ----------------------------------------------------------------
# 文章列表
# ----------------------------------------------------------------
def list_articles(column_id: int = 2, page: int = 1,
                  page_size: int = 50) -> list[dict]:
    """获取文章列表

    Args:
        column_id: 栏目ID（2=交易结果公告）
        page: 页码
        page_size: 每页条数

    Returns:
        [{iArticleid, sTitle, dPubDate, sAuthor, ...}, ...]
    """
    resp = _get_session().post(
        LIST_URL,
        json={
            "pageNo": page,
            "pageSize": page_size,
            "iColumnId": column_id,
            "channelCode": "04",
        },
        headers=HEADERS,
        timeout=15,
    )
    data = resp.json()
    if data.get("code") != "000000":
        raise RuntimeError(f"华南粮网列表接口失败: {data.get('msg')}")
    return data["result"]["orderBeanLists"]


def list_articles_by_month(year: int, month: int,
                           column_id: int = 2) -> list[dict]:
    """获取指定年月所有的文章

    遍历分页直到超出该月范围。
    """
    all_articles = []
    page = 1

    while True:
        articles = list_articles(column_id=column_id, page=page, page_size=50)
        if not articles:
            break

        stop = False
        for a in articles:
            pub = datetime.fromtimestamp(a["dPubDate"] / 1000, tz=CST)
            if pub.year == year and pub.month == month:
                all_articles.append(a)
            elif pub < datetime(year, month, 1, tzinfo=CST):
                stop = True
                break

        if stop:
            break
        page += 1
        if page > 50:
            break

    return all_articles


# ----------------------------------------------------------------
# 文章详情
# ----------------------------------------------------------------
def get_detail(article_id: int) -> dict:
    """获取文章详情

    Returns:
        {iArticleId, sTitle, sCentent (HTML), dPubDate, sAuthor, ...}
    """
    resp = _get_session().post(
        DETAIL_URL,
        json={"iArticleId": article_id, "channelCode": "04"},
        headers=HEADERS,
        timeout=15,
    )
    data = resp.json()
    if data.get("code") != "000000":
        raise RuntimeError(f"华南粮网详情接口失败: code={data.get('code')}, msg={data.get('msg')}, http_status={resp.status_code}")
    result = data.get("result")
    if not result:
        raise RuntimeError(f"华南粮网详情接口返回空数据 (articleId={article_id})")
    return result


# ----------------------------------------------------------------
# 附件处理
# ----------------------------------------------------------------
def extract_attachments(detail: dict) -> list[dict]:
    """从文章详情的 iAnnexList 字段中提取附件链接

    Returns:
        [{url, name, ext}, ...]
    """
    attrs = []
    for annex in detail.get("iAnnexList", []):
        url = annex.get("sAnnexUrl", "")
        name = annex.get("sAnnexName", "")
        ext = os.path.splitext(name)[1].lower()
        if not url:
            continue
        if not url.startswith("http"):
            url = f"{BASE_URL}{url}"
        attrs.append({"url": url, "name": name, "ext": ext})
    return attrs


def download_attachment(attach: dict, article_id: int) -> str | None:
    """下载附件到本地，返回本地文件路径"""
    os.makedirs(ATTACH_DIR, exist_ok=True)
    url = attach["url"]
    ext = attach["ext"]
    local_name = f"{article_id}_{attach['name']}"
    # 清理文件名中的非法字符
    local_name = re.sub(r'[\\/:*?"<>|]', "_", local_name)
    local_path = os.path.join(ATTACH_DIR, local_name)
    if not ext:
        # 从 URL 推断
        ext = os.path.splitext(url)[1]
        local_path += ext

    if os.path.exists(local_path):
        logger.debug("附件已存在: %s", local_path)
        return local_path

    try:
        resp = requests.get(url, headers=HEADERS, timeout=60)
        resp.raise_for_status()
        with open(local_path, "wb") as f:
            f.write(resp.content)
        logger.info("附件下载成功: %s (%d bytes)", local_path, len(resp.content))
        return local_path
    except Exception as e:
        logger.warning("附件下载失败 %s: %s", url, e)
        return None


# ----------------------------------------------------------------
# 表格文件读取
# ----------------------------------------------------------------
def spreadsheet_to_markdown(filepath: str) -> str | None:
    """将 .xls / .xlsx 文件内容转为 Markdown 表格

    Returns:
        连续的 Markdown 表格字符串（可能含多个 sheet），无法读取时返回 None
    """
    ext = os.path.splitext(filepath)[1].lower()
    rows_list = []  # [(sheet_name, [rows])]

    try:
        if ext == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    r = [str(c) if c is not None else "" for c in row]
                    # 跳过全空行
                    if any(cell.strip() for cell in r):
                        rows.append(r)
                if rows:
                    rows_list.append((sheet_name, rows))
            wb.close()

        elif ext == ".xls":
            import xlrd
            wb = xlrd.open_workbook(filepath)
            for i in range(wb.nsheets):
                ws = wb.sheet_by_index(i)
                sheet_name = ws.name
                rows = []
                for r in range(ws.nrows):
                    row_vals = [str(ws.cell_value(r, c)) for c in range(ws.ncols)]
                    if any(cell.strip() for cell in row_vals):
                        rows.append(row_vals)
                if rows:
                    rows_list.append((sheet_name, rows))
    except Exception as e:
        logger.warning("读取表格文件失败 %s: %s", filepath, e)
        return None

    if not rows_list:
        return None

    parts = []
    for sheet_name, rows in rows_list:
        if len(rows_list) > 1:
            parts.append(f"**{sheet_name}**")
            parts.append("")
        # 表头
        header = rows[0]
        parts.append("| " + " | ".join(header) + " |")
        parts.append("| " + " | ".join("---" for _ in header) + " |")
        # 数据行
        for row in rows[1:]:
            # 补齐列数
            while len(row) < len(header):
                row.append("")
            parts.append("| " + " | ".join(row[:len(header)]) + " |")
        parts.append("")

    return "\n".join(parts)


# ----------------------------------------------------------------
# 内容转换
# ----------------------------------------------------------------
_converter: html2text.HTML2Text | None = None


def _get_converter() -> html2text.HTML2Text:
    global _converter
    if _converter is None:
        h = html2text.HTML2Text()
        h.body_width = 0
        h.ignore_links = False
        h.ignore_images = False
        h.ignore_emphasis = False
        h.protect_links = True
        h.unicode_snob = True
        _converter = h
    return _converter


def build_markdown(detail: dict, local_attachments: list[str] = None) -> str:
    """构建文章 Markdown 内容"""
    title = detail.get("sTitle", "")
    pub_time = detail.get("dPubDate", "")
    author = detail.get("sAuthor", "")
    html_content = detail.get("sCentent", "")

    # HTML → Markdown
    body_md = _get_converter().handle(html_content).strip()

    # 组装
    parts = [
        f"# {title}",
        "",
        f"> **来源**：华南粮网",
        f"> **发布时间**：{pub_time}",
    ]
    if author:
        parts.append(f"> **作者**：{author}")
    parts.extend(["", "---", ""])

    if body_md:
        parts.extend(["## 正文", "", body_md, "", "---", ""])

    # 附件：表格文件直接嵌入内容，其他附件列文件名
    if local_attachments:
        spreadsheet_files = [f for f in local_attachments
                             if os.path.splitext(f)[1].lower() in (".xls", ".xlsx")]
        other_files = [f for f in local_attachments
                       if os.path.splitext(f)[1].lower() not in (".xls", ".xlsx")]

        if spreadsheet_files:
            parts.extend(["## 成交明细", ""])
            for f in spreadsheet_files:
                table_md = spreadsheet_to_markdown(f)
                if table_md:
                    parts.append(table_md)
                    parts.append("")
                else:
                    parts.append(f"- {os.path.basename(f)}（无法读取）")
                    parts.append("")

        if other_files:
            parts.extend(["## 附件", ""])
            for f in other_files:
                parts.append(f"- {os.path.basename(f)}")
            parts.append("")

    parts.append("*由数据采集系统自动同步至钉钉知识库*")

    return "\n".join(parts)
