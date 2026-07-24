"""国家粮食交易中心（grainmarket.com.cn）爬虫

采集指定区域市场的交易公告、交易清单、交易结果等信息。
"""
import logging
import os
import re
from datetime import datetime

import requests
import html2text

logger = logging.getLogger(__name__)

BASE_URL = "https://www.grainmarket.com.cn/centerweb"
API_URL = "https://www.grainmarket.com.cn/centerweb/getData"
ATTACH_DIR = "data/grainmarket_attachments"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
    "Content-Type": "application/x-www-form-urlencoded",
}

_SESSION: requests.Session | None = None


def _get_session() -> requests.Session:
    global _SESSION
    if _SESSION is None:
        _SESSION = requests.Session()
    return _SESSION


# ----------------------------------------------------------------
# 文章列表
# ----------------------------------------------------------------
# 类型映射
ARTICLE_TYPES = {
    "5": "交易公告",
    "8": "交易清单",
    "11": "交易结果",
    "2": "交易规则",
    "87": "公示通知",
}

# 同步采集的类型（交易规则不需要采集）
SYNC_TYPES = ["5", "8", "11", "87"]


def list_articles(market_id: str, article_type: str,
                  page: int = 1, page_size: int = 50) -> list[dict]:
    """获取指定市场和分类的文章列表

    Args:
        market_id: 市场 ID（如 S001031=海南）
        article_type: 文章类型（5=交易公告, 8=交易清单, 11=交易结果）
        page: 页码
        page_size: 每页条数

    Returns:
        [{articleID, title, publishtime, contentUrl, imgUrl, ...}, ...]
    """
    params = {
        "m": "tradeCenterOtherNewsList",
        "indexid": str(page),
        "pagesize": str(page_size),
        "articleTypeID": article_type,
        "marketId": market_id,
    }
    resp = _get_session().post(
        API_URL,
        data={"param": __import__("json").dumps(params)},
        headers=HEADERS,
        timeout=15,
    )
    data = resp.json()
    if data.get("code") != "001":
        raise RuntimeError(f"文章列表接口失败: {data.get('message')}")
    return data.get("data", [])


def list_articles_by_month(year: int, month: int,
                           market_id: str,
                           article_types: list[str] | None = None) -> list[dict]:
    """获取指定年月所有类型的文章

    Args:
        year: 年份
        month: 月份
        market_id: 市场 ID
        article_types: 文章类型列表，默认 SYNC_TYPES

    Returns:
        合并后的文章列表
    """
    all_articles = []
    seen_ids = set()
    article_types = article_types or SYNC_TYPES

    for at in article_types:
        page = 1
        while True:
            articles = list_articles(market_id, at, page=page, page_size=50)
            if not articles:
                break
            stop = False
            for a in articles:
                pub = a.get("publishtime", "")
                # 过滤指定年月
                prefix = f"{year:04d}-{month:02d}"
                if pub.startswith(prefix):
                    aid = a["articleID"]
                    if aid not in seen_ids:
                        seen_ids.add(aid)
                        a["_articleType"] = at
                        a["_articleTypeName"] = ARTICLE_TYPES.get(at, at)
                        all_articles.append(a)
                elif pub < prefix:
                    stop = True
                    break
            if stop:
                break
            page += 1
            if page > 50:
                break

    # 按发布时间排序
    all_articles.sort(key=lambda x: x.get("publishtime", ""), reverse=True)
    return all_articles


# ----------------------------------------------------------------
# 文章详情
# ----------------------------------------------------------------
def get_detail(article_id: str) -> dict:
    """获取文章详情

    Returns:
        {Title, Content(HTML), AddDate, ArticleFrom, MarketName, Annex, ...}
    """
    params = {
        "m": "tradeCenterNewsDetail",
        "articleId": article_id,
    }
    resp = _get_session().post(
        API_URL,
        data={"param": __import__("json").dumps(params)},
        headers=HEADERS,
        timeout=15,
    )
    data = resp.json()
    if data.get("code") != "001":
        raise RuntimeError(f"文章详情接口失败: {data.get('message')} (articleId={article_id})")
    result = data.get("data")
    if not result:
        raise RuntimeError(f"文章详情返回空数据 (articleId={article_id})")
    return result


# ----------------------------------------------------------------
# 附件处理
# ----------------------------------------------------------------
def extract_attachments(detail: dict) -> list[dict]:
    """从文章详情提取附件

    Annex 字段格式：可能是 URL 字符串或 null

    Returns:
        [{url, name, ext}, ...]
    """
    attrs = []
    annex = detail.get("Annex")
    if not annex:
        return attrs
    # Annex 可能是 list[{name, url}] 或逗号分隔的 URL 字符串
    if isinstance(annex, list):
        for item in annex:
            url = item.get("url", "")
            name = item.get("name", url.split("/")[-1] if "/" in url else url)
            ext = os.path.splitext(url.split("?")[0])[1].lower()
            if not url.startswith("http"):
                url = f"https://www.grainmarket.com.cn{url}" if url.startswith("/") else url
            attrs.append({"url": url, "name": name, "ext": ext})
    elif isinstance(annex, str):
        urls = [u.strip() for u in annex.split(",") if u.strip()]
        for url in urls:
            name = url.split("/")[-1] if "/" in url else url
            ext = os.path.splitext(name)[1].lower()
            if not url.startswith("http"):
                url = f"https://www.grainmarket.com.cn{url}" if url.startswith("/") else url
            attrs.append({"url": url, "name": name, "ext": ext})
    return attrs


def download_attachment(attach: dict, article_id: str) -> str | None:
    """下载附件到本地，返回本地文件路径"""
    os.makedirs(ATTACH_DIR, exist_ok=True)
    url = attach["url"]
    ext = attach["ext"]
    local_name = f"{article_id}_{attach['name']}"
    local_name = re.sub(r'[\\/:*?"<>|]', "_", local_name)
    local_path = os.path.join(ATTACH_DIR, local_name)

    if not ext:
        ext = os.path.splitext(url.split("?")[0])[1]
        local_path += ext

    if os.path.exists(local_path):
        logger.debug("附件已存在: %s", local_path)
        return local_path

    try:
        resp = requests.get(url, headers=HEADERS, timeout=60)
        resp.raise_for_status()
        with open(local_path, "wb") as f:
            f.write(resp.content)
        logger.info("附件下载成功: %s (%d bytes)", local_path, len(resp.content))
        return local_path
    except Exception as e:
        logger.warning("附件下载失败 %s: %s", url, e)
        return None


# ----------------------------------------------------------------
# 表格内容提取
# ----------------------------------------------------------------
def spreadsheet_to_markdown(filepath: str) -> str | None:
    """将 .xls / .xlsx 文件内容转为 Markdown 表格"""
    ext = os.path.splitext(filepath)[1].lower()
    rows_list = []

    try:
        if ext == ".xlsx":
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    r = [str(c) if c is not None else "" for c in row]
                    if any(cell.strip() for cell in r):
                        rows.append(r)
                if rows:
                    rows_list.append((sheet_name, rows))
            wb.close()

        elif ext == ".xls":
            import xlrd
            wb = xlrd.open_workbook(filepath)
            for i in range(wb.nsheets):
                ws = wb.sheet_by_index(i)
                sheet_name = ws.name
                rows = []
                for r in range(ws.nrows):
                    row_vals = [str(ws.cell_value(r, c)) for c in range(ws.ncols)]
                    if any(cell.strip() for cell in row_vals):
                        rows.append(row_vals)
                if rows:
                    rows_list.append((sheet_name, rows))
    except Exception as e:
        logger.warning("读取表格文件失败 %s: %s", filepath, e)
        return None

    if not rows_list:
        return None

    parts = []
    for sheet_name, rows in rows_list:
        if len(rows_list) > 1:
            parts.append(f"**{sheet_name}**")
            parts.append("")
        header = rows[0]
        parts.append("| " + " | ".join(header) + " |")
        parts.append("| " + " | ".join("---" for _ in header) + " |")
        for row in rows[1:]:
            while len(row) < len(header):
                row.append("")
            parts.append("| " + " | ".join(row[:len(header)]) + " |")
        parts.append("")

    return "\n".join(parts)


# ----------------------------------------------------------------
# 图片 OCR
# ----------------------------------------------------------------
def _download_content_image(detail: dict, article_id: str) -> str | None:
    """下载文章内容中的图片到本地，返回本地路径"""
    html_content = detail.get("Content", "")
    urls = re.findall(r'<img[^>]+src="([^"]+)"', html_content)
    if not urls:
        return None
    url = urls[0]
    if not url.startswith("http"):
        url = f"https://www.grainmarket.com.cn{url}" if url.startswith("/") else url
    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        os.makedirs(ATTACH_DIR, exist_ok=True)
        ext = os.path.splitext(url.split("?")[0])[1] or ".png"
        local_path = os.path.join(ATTACH_DIR, f"{article_id}_content{ext}")
        with open(local_path, "wb") as f:
            f.write(resp.content)
        logger.info("内容图片已下载: %s (%d bytes)", local_path, len(resp.content))
        return local_path
    except Exception as e:
        logger.warning("内容图片下载失败 %s: %s", url, e)
        return None


# ----------------------------------------------------------------
# 内容转换
# ----------------------------------------------------------------
_converter: html2text.HTML2Text | None = None


def _get_converter() -> html2text.HTML2Text:
    global _converter
    if _converter is None:
        h = html2text.HTML2Text()
        h.body_width = 0
        h.ignore_links = False
        h.ignore_images = False
        h.ignore_emphasis = False
        h.protect_links = True
        h.unicode_snob = True
        _converter = h
    return _converter


def build_markdown(detail: dict, local_attachments: list[str] = None,
                   article_type_name: str = "", article_id: str = "") -> str:
    """构建文章 Markdown 内容"""
    title = detail.get("Title", "")
    add_date = detail.get("AddDate", "")
    article_from = detail.get("ArticleFrom", "")
    market_name = detail.get("MarketName", "")
    html_content = detail.get("Content", "")

    # HTML → Markdown
    body_md = _get_converter().handle(html_content).strip()

    parts = [
        f"# {title}",
        "",
        f"> **来源**：{article_from}",
        f"> **市场**：{market_name}",
        f"> **发布时间**：{add_date}",
    ]
    if article_type_name:
        parts.append(f"> **类型**：{article_type_name}")
    parts.extend(["", "---", ""])

    if body_md:
        # 判断内容是否为纯图片
        is_image_only = bool(re.search(r'<img[^>]+src="[^"]+"', html_content))
        if is_image_only:
            # 先保留原图
            parts.extend(["## 内容（原图）", "", body_md, "", "---", ""])
            # 尝试 OCR 识别图片中的文字
            ocr_text = None
            if article_id:
                img_path = _download_content_image(detail, article_id)
                if img_path:
                    try:
                        from utils.ocr import recognize_table
                        ocr_text = recognize_table(img_path)
                        logger.info("OCR 识别成功: %s (%d 字符)", article_id, len(ocr_text))
                    except Exception as e:
                        logger.warning("OCR 识别失败 %s: %s", article_id, e)

            if ocr_text:
                if "\n---\n" in ocr_text[:500] or "吨" in ocr_text[:200]:
                    parts.extend(["## 交易记录", "", ocr_text, "", "---", ""])
                else:
                    parts.extend(["## OCR 识别结果", "", "```", ocr_text, "```", "", "---", ""])
        else:
            parts.extend(["## 正文", "", body_md, "", "---", ""])

    # 附件：表格文件直接嵌入，其他列文件名
    if local_attachments:
        spreadsheet_files = [f for f in local_attachments
                             if os.path.splitext(f)[1].lower() in (".xls", ".xlsx")]
        other_files = [f for f in local_attachments
                       if os.path.splitext(f)[1].lower() not in (".xls", ".xlsx")]

        if spreadsheet_files:
            parts.extend(["## 成交明细", ""])
            for f in spreadsheet_files:
                table_md = spreadsheet_to_markdown(f)
                if table_md:
                    parts.append(table_md)
                    parts.append("")
                else:
                    parts.append(f"- {os.path.basename(f)}（无法读取）")
                    parts.append("")

        if other_files:
            parts.extend(["## 附件", ""])
            for f in other_files:
                parts.append(f"- {os.path.basename(f)}")
            parts.append("")

    parts.append("*由数据采集系统自动同步至钉钉知识库*")

    return "\n".join(parts)
